"""
Export Service Module - Refactored

Provides a facade for Excel and CSV export functionality.
Sheet creation logic has been moved to specialized modules.
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Any, TYPE_CHECKING, Optional

if TYPE_CHECKING:
    from services.device_service import DeviceService
from openpyxl import Workbook

from services.export.sheets import (
    SummarySheet,
    DeviceSheet,
    RiskSheet,
    ComplianceSheet,
    CVESummarySheet,
    CVEDetailsSheet,
    CVETopSheet,
)
from services.export.base import get_connection_status_info
from utils.device_filters import filter_devices_for_export
from services.risk_service import RiskService

logger = logging.getLogger(__name__)


class ExportService:
    """Facade for handling device data exports"""
    
    def __init__(self, device_service: 'DeviceService'):
        self.device_service = device_service
    
    def export_devices_to_excel(self, filter_args: Optional[Dict[str, Any]] = None, 
                                devices: Optional[List[Dict]] = None) -> str:
        """Export filtered devices to Excel file with multiple sheets"""
        logger.info(f"Starting Excel export with devices={devices is not None}, filter_args={filter_args}")

        if devices is None:
            devices = self._get_devices_for_export()

        if filter_args:
            filtered_devices = filter_devices_for_export(devices, filter_args)
        else:
            filtered_devices = devices

        logger.info(f"After filtering: {len(filtered_devices)} devices to export")

        if not filtered_devices:
            raise Exception("No data to export")

        filepath = self._create_device_workbook(filtered_devices)
        logger.info(f"Exported {len(filtered_devices)} devices to Excel: {filepath}")
        return filepath
    
    def _get_devices_for_export(self) -> List[Dict]:
        """Get device data for export from API or sample data"""
        config = self.device_service.config
        
        if config.USE_SAMPLE_DATA:
            return self.device_service._load_sample_data()
        
        client = self.device_service.get_lookout_client()
        if client is None:
            raise Exception("API client not available")
        
        if not client.is_authenticated():
            client.authenticate()
        
        logger.info("Fetching devices from Lookout API for export")
        raw_devices = client.get_devices(limit=1000)
        
        from device_cache import enhanced_device_mapping
        devices = [enhanced_device_mapping(device) for device in raw_devices]
        logger.info(f"Successfully fetched {len(devices)} devices from API")
        
        return devices
    
    def _create_device_workbook(self, devices: List[Dict]) -> str:
        """Create Excel workbook with device data using sheet modules"""
        logger.info(f"Creating Excel workbook for {len(devices)} devices")

        wb = Workbook()
        
        SummarySheet.create(wb, devices)
        DeviceSheet.create(wb, devices)
        RiskSheet.create(wb, devices)
        ComplianceSheet.create(wb, devices)
        
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        filename = f"device_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('/tmp', filename)
        wb.save(filepath)

        logger.info(f"Excel workbook saved to: {filepath}")
        return filepath
    
    def create_csv_export(self, devices: List[Dict]) -> str:
        """Create CSV export of device data"""
        headers = [
            'Device Name', 'User Email', 'Platform', 'Risk Level',
            'Active Issues', 'Last Check-in', 'OS Version', 'Compliance Status',
            'Device Group', 'MDM ID', 'Connection Status', 'Threat Family Names', 'Threat Descriptions'
        ]
        
        rows = []
        for device in devices:
            risk_analysis = RiskService.analyze_device_risk(device)
            active_issues_count = risk_analysis.get('total_issues', 0)
            
            days_since = device.get('days_since_checkin', -1)
            connection_info = get_connection_status_info(days_since)

            row = [
                device.get('device_name', ''),
                device.get('user_email', ''),
                device.get('platform', ''),
                device.get('risk_level', ''),
                str(active_issues_count),
                device.get('last_checkin', ''),
                device.get('os_version', ''),
                device.get('compliance_status', ''),
                device.get('device_group_name', 'N/A'),
                device.get('mdm_id', 'N/A'),
                connection_info['label'],
                ', '.join(device.get('threat_family_names', [])),
                ', '.join(device.get('threat_descriptions', []))
            ]
            rows.append(row)
        
        csv_rows = [headers] + rows
        csv_content = '\n'.join([
            ','.join([f'"{str(field).replace(chr(34), chr(34)+chr(34))}"' for field in row])
            for row in csv_rows
        ])
        
        return csv_content
    
    def export_cve_report_to_excel(self, cve_report: Dict[str, Any]) -> str:
        """Export CVE vulnerability report to Excel"""
        logger.info("Creating CVE vulnerability Excel report")
        
        wb = Workbook()
        
        CVESummarySheet.create(wb, cve_report)
        CVEDetailsSheet.create(wb, cve_report)
        CVETopSheet.create(wb, cve_report)
        
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        filename = f"cve_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('/tmp', filename)
        wb.save(filepath)
        
        logger.info(f"CVE report saved to: {filepath}")
        return filepath
    
    def get_export_stats(self, devices: List[Dict]) -> Dict[str, Any]:
        """Get statistics about devices being exported"""
        total_devices = len(devices)
        
        risk_counts = {}
        platform_counts = {}
        
        for device in devices:
            risk_level = device.get('risk_level', 'Unknown')
            platform = device.get('platform', 'Unknown')
            
            risk_counts[risk_level] = risk_counts.get(risk_level, 0) + 1
            platform_counts[platform] = platform_counts.get(platform, 0) + 1
        
        return {
            'total_devices': total_devices,
            'risk_level_distribution': risk_counts,
            'platform_distribution': platform_counts,
            'export_timestamp': datetime.now().isoformat()
        }
