"""
Base Excel utilities for export services.

Shared formatting utilities and base classes for Excel sheet creation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Any


class ExcelStyles:
    """Shared Excel styling utilities"""
    
    # Header styles
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    # Title styles
    TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
    TITLE_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Risk level colors
    CRITICAL_FILL = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    HIGH_FILL = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
    MEDIUM_FILL = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    LOW_FILL = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    SECURE_FILL = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    
    # Section header
    SECTION_FONT = Font(bold=True, size=12)
    SECTION_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    @classmethod
    def apply_risk_color(cls, cell, risk_level: str):
        """Apply color coding based on risk level"""
        risk_level = risk_level.lower()
        if risk_level == 'critical':
            cell.fill = cls.CRITICAL_FILL
            cell.font = Font(color="FFFFFF", bold=True)
        elif risk_level == 'high':
            cell.fill = cls.HIGH_FILL
        elif risk_level == 'medium':
            cell.fill = cls.MEDIUM_FILL
        elif risk_level == 'low':
            cell.fill = cls.LOW_FILL
            cell.font = Font(color="FFFFFF")
        elif risk_level == 'secure':
            cell.fill = cls.SECURE_FILL
            cell.font = Font(color="FFFFFF")
    
    @classmethod
    def apply_severity_color(cls, cell, severity: str):
        """Apply color coding based on severity level"""
        severity = severity.lower()
        if severity == 'critical':
            cell.fill = cls.CRITICAL_FILL
            cell.font = Font(color="FFFFFF", bold=True)
        elif severity == 'high':
            cell.fill = cls.HIGH_FILL
        elif severity == 'medium':
            cell.fill = cls.MEDIUM_FILL
        elif severity == 'low':
            cell.fill = cls.LOW_FILL
            cell.font = Font(color="FFFFFF")


def get_connection_status_info(days_since: int) -> Dict[str, str]:
    """Get connection status information for export"""
    if days_since == -1:
        return {'status': 'never_connected', 'label': 'Never Connected'}
    elif days_since <= 1:
        return {'status': 'connected', 'label': 'Connected'}
    elif days_since <= 7:
        return {'status': 'recent', 'label': 'Recent'}
    elif days_since <= 30:
        return {'status': 'stale', 'label': 'Stale'}
    elif days_since <= 90:
        return {'status': 'disconnected', 'label': 'Disconnected'}
    else:
        return {'status': 'very_stale', 'label': 'Very Stale'}
