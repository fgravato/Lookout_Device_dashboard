"""
CVE Summary sheet for vulnerability reports.
"""

import logging
from datetime import datetime
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.export.base import ExcelStyles

logger = logging.getLogger(__name__)


class CVESummarySheet:
    """Creates the CVE summary sheet"""
    
    @staticmethod
    def create(wb: Workbook, report: Dict[str, Any]):
        """Create CVE summary sheet"""
        ws = wb.create_sheet("CVE Summary", 0)
        
        ws['A1'] = 'CVE Vulnerability Report - Summary'
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        summary = report.get('summary', {})
        
        row = 3
        ws[f'A{row}'] = 'Report Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws[f'A{row}'] = 'Total Devices Scanned:'
        ws[f'B{row}'] = summary.get('total_devices', 0)
        row += 1
        ws[f'A{row}'] = 'Devices with Vulnerabilities:'
        ws[f'B{row}'] = summary.get('devices_with_vulnerabilities', 0)
        row += 1
        ws[f'A{row}'] = 'Vulnerability Percentage:'
        ws[f'B{row}'] = f"{summary.get('vulnerability_percentage', 0)}%"
        row += 1
        ws[f'A{row}'] = 'Total Unique CVEs Found:'
        ws[f'B{row}'] = summary.get('total_cves_found', 0)
        
        row += 2
        ws[f'A{row}'] = 'Severity Breakdown'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = 'Severity Level'
        ws[f'B{row}'] = 'CVE Count'
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = Font(bold=True)
            cell.fill = ExcelStyles.SECTION_FILL
        row += 1
        
        severity_breakdown = summary.get('severity_breakdown', {})
        for severity in ['Critical', 'High', 'Medium', 'Low']:
            count = severity_breakdown.get(severity, 0)
            if count > 0:
                ws[f'A{row}'] = severity
                ws[f'B{row}'] = count
                ExcelStyles.apply_severity_color(ws[f'A{row}'], severity)
                row += 1
        
        row += 1
        metadata = report.get('scan_metadata', {})
        ws[f'A{row}'] = 'Scan Details'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Android Patches Scanned:'
        ws[f'B{row}'] = metadata.get('android_patches_scanned', 0)
        row += 1
        ws[f'A{row}'] = 'iOS Versions Scanned:'
        ws[f'B{row}'] = metadata.get('ios_versions_scanned', 0)
        row += 1
        ws[f'A{row}'] = 'Minimum Severity Threshold:'
        ws[f'B{row}'] = metadata.get('minimum_severity', 7)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
