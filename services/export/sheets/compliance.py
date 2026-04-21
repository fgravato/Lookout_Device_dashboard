"""
Compliance tracking sheet for device exports.
"""

import logging
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.export.base import ExcelStyles

logger = logging.getLogger(__name__)


class ComplianceSheet:
    """Creates the compliance tracking sheet"""
    
    @staticmethod
    def create(wb: Workbook, devices: List[Dict]):
        """Create compliance tracking sheet"""
        ws = wb.create_sheet("Compliance Status")
        
        ws['A1'] = 'Compliance Status Report'
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:E1')
        
        row = 3
        
        compliance_groups = {}
        for device in devices:
            status = device.get('compliance_status', 'Unknown')
            if status not in compliance_groups:
                compliance_groups[status] = []
            compliance_groups[status].append(device)
        
        ws[f'A{row}'] = 'Compliance Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for status, device_list in sorted(compliance_groups.items()):
            ws[f'A{row}'] = status
            ws[f'B{row}'] = len(device_list)
            ws[f'C{row}'] = f"{(len(device_list)/len(devices)*100):.1f}%"
            row += 1
        
        row += 1
        
        ws[f'A{row}'] = 'Device Details'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Device Name', 'User Email', 'Platform', 'Compliance Status', 'Protection Status', 'Activation Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
        
        row += 1
        
        for device in sorted(devices, key=lambda d: d.get('compliance_status', 'Unknown')):
            ws.cell(row=row, column=1, value=device.get('device_name', 'Unknown'))
            ws.cell(row=row, column=2, value=device.get('user_email', 'N/A'))
            ws.cell(row=row, column=3, value=device.get('platform', 'Unknown'))
            ws.cell(row=row, column=4, value=device.get('compliance_status', 'Unknown'))
            ws.cell(row=row, column=5, value=device.get('protection_status', 'Unknown'))
            ws.cell(row=row, column=6, value=device.get('activation_status', 'Unknown'))
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        ws.freeze_panes = 'A9'
