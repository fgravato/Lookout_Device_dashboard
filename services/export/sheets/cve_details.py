"""
CVE Details sheet for vulnerability reports.
"""

import logging
from typing import Dict, Any, List

from openpyxl import Workbook

from services.export.base import ExcelStyles

logger = logging.getLogger(__name__)


class CVEDetailsSheet:
    """Creates the detailed CVE list sheet"""
    
    @staticmethod
    def create(wb: Workbook, report: Dict[str, Any]):
        """Create detailed CVE list sheet"""
        ws = wb.create_sheet("All CVEs")
        
        headers = ['CVE ID', 'Severity Score', 'Severity Level', 'Platform', 
                   'OS Version/Patch', 'Affected Devices', 'Description']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
        
        vulnerabilities = report.get('all_vulnerabilities', [])
        
        vulnerabilities.sort(key=lambda v: (v.get('severity', 0), v.get('affected_device_count', 0)), reverse=True)
        
        for row_idx, vuln in enumerate(vulnerabilities, 2):
            severity = vuln.get('severity', 0)
            severity_label = vuln.get('severity_label', 'Unknown')
            
            ws.cell(row=row_idx, column=1, value=vuln.get('cve', 'Unknown'))
            ws.cell(row=row_idx, column=2, value=severity)
            
            severity_cell = ws.cell(row=row_idx, column=3, value=severity_label)
            ExcelStyles.apply_severity_color(severity_cell, severity_label)
            
            ws.cell(row=row_idx, column=4, value=vuln.get('platform', 'Unknown'))
            ws.cell(row=row_idx, column=5, value=vuln.get('patch_level') or vuln.get('os_version', 'Unknown'))
            ws.cell(row=row_idx, column=6, value=vuln.get('affected_device_count', 0))
            ws.cell(row=row_idx, column=7, value=vuln.get('description', 'No description'))
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:G{len(vulnerabilities)+1}"
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 60
