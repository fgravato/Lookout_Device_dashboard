"""
Top CVEs sheet for vulnerability reports.
"""

import logging
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.export.base import ExcelStyles

logger = logging.getLogger(__name__)


class CVETopSheet:
    """Creates the top CVEs by affected devices sheet"""
    
    @staticmethod
    def create(wb: Workbook, report: Dict[str, Any]):
        """Create top CVEs sheet"""
        ws = wb.create_sheet("Top CVEs")
        
        ws['A1'] = 'Top 10 CVEs by Affected Device Count'
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        ws.merge_cells('A1:C1')
        
        row = 3
        headers = ['Rank', 'CVE ID', 'Affected Devices']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = ExcelStyles.SECTION_FILL
        
        row += 1
        top_cves = report.get('top_cves', [])[:10]
        
        for rank, cve_data in enumerate(top_cves, 1):
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=cve_data.get('cve', 'Unknown'))
            ws.cell(row=row, column=3, value=cve_data.get('affected_devices', 0))
            
            if rank <= 3:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            row += 1
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
