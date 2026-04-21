"""
Risk analysis sheet for device exports.
"""

import logging
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.export.base import ExcelStyles
from services.risk_service import RiskService

logger = logging.getLogger(__name__)


class RiskSheet:
    """Creates the risk analysis sheet with devices grouped by issues"""
    
    @staticmethod
    def create(wb: Workbook, devices: List[Dict]):
        """Create risk analysis sheet"""
        ws = wb.create_sheet("Risk Analysis")
        
        ws['A1'] = 'Risk Analysis - Devices with Active Issues'
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        row = 3
        devices_with_issues = []
        
        for device in devices:
            risk_analysis = RiskService.analyze_device_risk(device)
            if risk_analysis.get('total_issues', 0) > 0:
                devices_with_issues.append({
                    'device': device,
                    'analysis': risk_analysis
                })
        
        ws[f'A{row}'] = f'Devices with Issues: {len(devices_with_issues)} of {len(devices)}'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        headers = ['Device Name', 'User Email', 'Risk Level', 'Total Issues', 'Issue Categories', 'Recommendations']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
        
        row += 1
        
        devices_with_issues.sort(key=lambda x: x['analysis']['total_issues'], reverse=True)
        
        for item in devices_with_issues:
            device = item['device']
            analysis = item['analysis']
            
            ws.cell(row=row, column=1, value=device.get('device_name', 'Unknown'))
            ws.cell(row=row, column=2, value=device.get('user_email', 'N/A'))
            ws.cell(row=row, column=3, value=device.get('risk_level', 'Unknown'))
            ws.cell(row=row, column=4, value=analysis['total_issues'])
            
            categories = ', '.join(set([f['category'] for f in analysis.get('risk_factors', [])]))
            ws.cell(row=row, column=5, value=categories)
            
            recommendations = '; '.join(analysis.get('recommendations', [])[:3])
            ws.cell(row=row, column=6, value=recommendations)
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 50
        
        ws.freeze_panes = 'A6'
