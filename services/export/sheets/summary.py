"""
Executive Summary Sheet for device exports.
"""

import logging
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.export.base import ExcelStyles, get_connection_status_info
from services.risk_service import RiskService

logger = logging.getLogger(__name__)


class SummarySheet:
    """Creates the Executive Summary sheet with key metrics"""
    
    @staticmethod
    def create(wb: Workbook, devices: List[Dict]):
        """Create executive summary sheet with key metrics"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = 'Fleet Management Report - Executive Summary'
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Report metadata
        ws['A3'] = 'Report Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Total Devices:'
        ws['B4'] = len(devices)
        
        # Calculate statistics
        platform_counts = {}
        risk_counts = {}
        compliance_counts = {}
        connection_counts = {'connected': 0, 'recent': 0, 'stale': 0, 'disconnected': 0, 'very_stale': 0}
        mdm_counts = {}
        tenant_counts = {}
        total_issues = 0
        
        for device in devices:
            platform = device.get('platform', 'Unknown')
            risk = device.get('risk_level', 'Unknown')
            compliance = device.get('compliance_status', 'Unknown')
            
            platform_counts[platform] = platform_counts.get(platform, 0) + 1
            risk_counts[risk] = risk_counts.get(risk, 0) + 1
            compliance_counts[compliance] = compliance_counts.get(compliance, 0) + 1
            
            # Multi-tenant statistics
            mdm_provider = device.get('mdm_provider', 'N/A')
            tenant_name = device.get('tenant_name', 'N/A')
            mdm_counts[mdm_provider] = mdm_counts.get(mdm_provider, 0) + 1
            tenant_counts[tenant_name] = tenant_counts.get(tenant_name, 0) + 1
            
            # Calculate connection status
            try:
                last_checkin = datetime.fromisoformat((device['last_checkin'] or '').replace('Z', '+00:00'))
                days_since = (datetime.now().replace(tzinfo=last_checkin.tzinfo) - last_checkin).days
            except (ValueError, KeyError, TypeError):
                days_since = -1
            
            if days_since <= 1:
                connection_counts['connected'] += 1
            elif days_since <= 7:
                connection_counts['recent'] += 1
            elif days_since <= 30:
                connection_counts['stale'] += 1
            elif days_since <= 90:
                connection_counts['disconnected'] += 1
            else:
                connection_counts['very_stale'] += 1
            
            risk_analysis = RiskService.analyze_device_risk(device)
            total_issues += risk_analysis.get('total_issues', 0)
        
        # Platform breakdown
        row = 6
        ws[f'A{row}'] = 'Platform Distribution'
        ws[f'A{row}'].font = ExcelStyles.SECTION_FONT
        row += 1
        ws[f'A{row}'] = 'Platform'
        ws[f'B{row}'] = 'Count'
        ws[f'C{row}'] = 'Percentage'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
            cell.fill = ExcelStyles.SECTION_FILL
        
        row += 1
        for platform, count in sorted(platform_counts.items()):
            ws[f'A{row}'] = platform
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count/len(devices)*100):.1f}%"
            row += 1
        
        # Risk level breakdown
        row += 1
        ws[f'A{row}'] = 'Risk Level Distribution'
        ws[f'A{row}'].font = ExcelStyles.SECTION_FONT
        row += 1
        ws[f'A{row}'] = 'Risk Level'
        ws[f'B{row}'] = 'Count'
        ws[f'C{row}'] = 'Percentage'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
            cell.fill = ExcelStyles.SECTION_FILL
        
        row += 1
        risk_order = ['Critical', 'High', 'Medium', 'Low', 'Secure']
        for risk_level in risk_order:
            count = risk_counts.get(risk_level, 0)
            if count > 0:
                ws[f'A{row}'] = risk_level
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{(count/len(devices)*100):.1f}%"
                ExcelStyles.apply_risk_color(ws[f'A{row}'], risk_level)
                row += 1
        
        # Connection status
        row += 1
        ws[f'A{row}'] = 'Connection Status'
        ws[f'A{row}'].font = ExcelStyles.SECTION_FONT
        row += 1
        ws[f'A{row}'] = 'Status'
        ws[f'B{row}'] = 'Count'
        ws[f'C{row}'] = 'Percentage'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
            cell.fill = ExcelStyles.SECTION_FILL
        
        row += 1
        for status, count in connection_counts.items():
            if count > 0:
                ws[f'A{row}'] = status.replace('_', ' ').title()
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{(count/len(devices)*100):.1f}%"
                row += 1
        
        # Key metrics
        row += 1
        ws[f'A{row}'] = 'Key Metrics'
        ws[f'A{row}'].font = ExcelStyles.SECTION_FONT
        row += 1
        ws[f'A{row}'] = 'Total Active Issues:'
        ws[f'B{row}'] = total_issues
        row += 1
        ws[f'A{row}'] = 'Average Issues per Device:'
        ws[f'B{row}'] = f"{(total_issues/len(devices)):.2f}"
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
