"""
Detailed device listing sheet for Excel exports.
"""

import logging
from typing import List, Dict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from services.export.base import ExcelStyles, get_connection_status_info
from services.risk_service import RiskService

logger = logging.getLogger(__name__)


class DeviceSheet:
    """Creates the detailed device listing sheet"""
    
    @staticmethod
    def create(wb: Workbook, devices: List[Dict]):
        """Create detailed device listing sheet"""
        ws = wb.create_sheet("All Devices")

        headers = [
            'Device Name', 'Device ID', 'User Email', 'Platform', 'Risk Level',
            'Active Issues', 'Last Check-in', 'Days Since Check-in', 'OS Version',
            'Security Patch', 'App Version', 'Compliance Status', 'Connection Status',
            'Device Group', 'MDM ID', 'Manufacturer', 'Model', 'Activation Status',
            'Threat Family Names', 'Threat Descriptions',
            'MDM Provider', 'MDM Identifier', 'Tenant ID', 'Tenant Name'
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row, device in enumerate(devices, 2):
            try:
                last_checkin = datetime.fromisoformat((device['last_checkin'] or '').replace('Z', '+00:00'))
                days_since = (datetime.now().replace(tzinfo=last_checkin.tzinfo) - last_checkin).days
            except (ValueError, KeyError, TypeError):
                days_since = -1

            risk_analysis = RiskService.analyze_device_risk(device)
            active_issues_count = risk_analysis.get('total_issues', 0)
            connection_info = get_connection_status_info(days_since)
            risk_level = device.get('risk_level', 'Unknown')

            ws.cell(row=row, column=1, value=device.get('device_name', 'Unknown'))
            ws.cell(row=row, column=2, value=device.get('device_id', ''))
            ws.cell(row=row, column=3, value=device.get('user_email', 'N/A'))
            ws.cell(row=row, column=4, value=device.get('platform', 'Unknown'))
            
            risk_cell = ws.cell(row=row, column=5, value=risk_level)
            ExcelStyles.apply_risk_color(risk_cell, risk_level)
            
            ws.cell(row=row, column=6, value=active_issues_count)
            ws.cell(row=row, column=7, value=device.get('last_checkin', ''))
            ws.cell(row=row, column=8, value=days_since)
            ws.cell(row=row, column=9, value=device.get('os_version', 'Unknown'))
            ws.cell(row=row, column=10, value=device.get('security_patch_level', 'N/A'))
            ws.cell(row=row, column=11, value=device.get('app_version', 'Unknown'))
            ws.cell(row=row, column=12, value=device.get('compliance_status', 'Unknown'))
            ws.cell(row=row, column=13, value=connection_info['label'])
            ws.cell(row=row, column=14, value=device.get('device_group_name', 'N/A'))
            ws.cell(row=row, column=15, value=device.get('mdm_id', 'N/A'))
            ws.cell(row=row, column=16, value=device.get('manufacturer', 'N/A'))
            ws.cell(row=row, column=17, value=device.get('model', 'N/A'))
            ws.cell(row=row, column=18, value=device.get('activation_status', 'Unknown'))
            ws.cell(row=row, column=19, value=', '.join(device.get('threat_family_names', [])))
            ws.cell(row=row, column=20, value=', '.join(device.get('threat_descriptions', [])))
            ws.cell(row=row, column=21, value=device.get('mdm_connector_id', 'N/A'))
            ws.cell(row=row, column=22, value=device.get('mdm_connector_uuid', 'N/A'))
            ws.cell(row=row, column=23, value=device.get('external_id', 'N/A'))
            ws.cell(row=row, column=24, value=device.get('mdm_provider', 'N/A'))
            ws.cell(row=row, column=25, value=device.get('mdm_identifier', 'N/A'))
            ws.cell(row=row, column=26, value=device.get('tenant_id', 'N/A'))
            ws.cell(row=row, column=27, value=device.get('tenant_name', 'N/A'))

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}1"
        
        for col_idx, column in enumerate(ws.iter_cols(min_col=1, max_col=len(headers)), 1):
            max_length = len(str(headers[col_idx-1]))
            for cell in list(column)[1:min(101, len(devices)+1)]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
