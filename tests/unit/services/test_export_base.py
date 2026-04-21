"""
Tests for the export base utilities.
"""

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.export.base import ExcelStyles, get_connection_status_info


class TestExcelStyles:
    """Test suite for ExcelStyles utility class"""
    
    def test_header_font_properties(self):
        """Test header font has correct properties"""
        assert ExcelStyles.HEADER_FONT.bold == True
        assert ExcelStyles.HEADER_FONT.color.rgb == '00FFFFFF'  # White
    
    def test_header_fill_properties(self):
        """Test header fill has correct color"""
        assert ExcelStyles.HEADER_FILL.start_color.rgb == '00366092'  # Blue
        assert ExcelStyles.HEADER_FILL.fill_type == 'solid'
    
    def test_risk_colors_defined(self):
        """Test all risk level colors are defined"""
        assert ExcelStyles.CRITICAL_FILL.start_color.rgb == '00DC3545'  # Red
        assert ExcelStyles.HIGH_FILL.start_color.rgb == '00FD7E14'      # Orange
        assert ExcelStyles.MEDIUM_FILL.start_color.rgb == '00FFC107'    # Yellow
        assert ExcelStyles.LOW_FILL.start_color.rgb == '0017A2B8'       # Blue
        assert ExcelStyles.SECURE_FILL.start_color.rgb == '0028A745'    # Green


class TestApplyRiskColor:
    """Test suite for apply_risk_color method"""
    
    def test_apply_critical_risk_color(self):
        """Test Critical risk level gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_risk_color(cell, 'Critical')
        
        assert cell.fill.start_color.rgb == '00DC3545'
        assert cell.font.color.rgb == '00FFFFFF'
        assert cell.font.bold == True
    
    def test_apply_high_risk_color(self):
        """Test High risk level gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_risk_color(cell, 'High')
        
        assert cell.fill.start_color.rgb == '00FD7E14'
    
    def test_apply_medium_risk_color(self):
        """Test Medium risk level gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_risk_color(cell, 'Medium')
        
        assert cell.fill.start_color.rgb == '00FFC107'
    
    def test_apply_low_risk_color(self):
        """Test Low risk level gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_risk_color(cell, 'Low')
        
        assert cell.fill.start_color.rgb == '0017A2B8'
        assert cell.font.color.rgb == '00FFFFFF'
    
    def test_apply_secure_risk_color(self):
        """Test Secure risk level gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_risk_color(cell, 'Secure')
        
        assert cell.fill.start_color.rgb == '0028A745'
        assert cell.font.color.rgb == '00FFFFFF'
    
    def test_apply_risk_color_case_insensitive(self):
        """Test risk color application is case insensitive"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_risk_color(cell, 'critical')  # lowercase
        
        assert cell.fill.start_color.rgb == '00DC3545'


class TestApplySeverityColor:
    """Test suite for apply_severity_color method"""
    
    def test_apply_critical_severity(self):
        """Test Critical severity gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_severity_color(cell, 'Critical')
        
        assert cell.fill.start_color.rgb == '00DC3545'
        assert cell.font.bold == True
    
    def test_apply_high_severity(self):
        """Test High severity gets correct styling"""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        
        ExcelStyles.apply_severity_color(cell, 'High')
        
        assert cell.fill.start_color.rgb == '00FD7E14'


class TestGetConnectionStatusInfo:
    """Test suite for get_connection_status_info function"""
    
    def test_never_connected_status(self):
        """Test -1 days returns never_connected status"""
        result = get_connection_status_info(-1)
        assert result['status'] == 'never_connected'
        assert result['label'] == 'Never Connected'
    
    def test_connected_status(self):
        """Test 0-1 days returns connected status"""
        result = get_connection_status_info(0)
        assert result['status'] == 'connected'
        assert result['label'] == 'Connected'
        
        result = get_connection_status_info(1)
        assert result['status'] == 'connected'
    
    def test_recent_status(self):
        """Test 2-7 days returns recent status"""
        result = get_connection_status_info(2)
        assert result['status'] == 'recent'
        assert result['label'] == 'Recent'
        
        result = get_connection_status_info(7)
        assert result['status'] == 'recent'
    
    def test_stale_status(self):
        """Test 8-30 days returns stale status"""
        result = get_connection_status_info(8)
        assert result['status'] == 'stale'
        assert result['label'] == 'Stale'
        
        result = get_connection_status_info(30)
        assert result['status'] == 'stale'
    
    def test_disconnected_status(self):
        """Test 31-90 days returns disconnected status"""
        result = get_connection_status_info(31)
        assert result['status'] == 'disconnected'
        assert result['label'] == 'Disconnected'
        
        result = get_connection_status_info(90)
        assert result['status'] == 'disconnected'
    
    def test_very_stale_status(self):
        """Test >90 days returns very_stale status"""
        result = get_connection_status_info(91)
        assert result['status'] == 'very_stale'
        assert result['label'] == 'Very Stale'
        
        result = get_connection_status_info(365)
        assert result['status'] == 'very_stale'
